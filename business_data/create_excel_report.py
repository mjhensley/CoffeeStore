"""
Generate formatted Excel report for Grainhouse Coffee business data
"""

import json
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call(['pip', 'install', 'openpyxl'])
    import openpyxl
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

# Load the JSON data
with open('purchase_orders_and_tracking.json', 'r') as f:
    data = json.load(f)

# Create workbook
wb = openpyxl.Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4A3728", end_color="4A3728", fill_type="solid")
subheader_fill = PatternFill(start_color="C9A96E", end_color="C9A96E", fill_type="solid")
subheader_font = Font(bold=True, color="1A1A1A", size=10)
currency_format = '"$"#,##0.00'
date_format = 'MMM D, YYYY'
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
delivered_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
transit_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
pending_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

def apply_header_style(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def apply_subheader_style(cell):
    cell.font = subheader_font
    cell.fill = subheader_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def auto_fit_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# ============================================
# SHEET 1: Dashboard / Summary
# ============================================
ws_summary = wb.active
ws_summary.title = "Dashboard"

# Title
ws_summary.merge_cells('A1:F1')
ws_summary['A1'] = "GRAINHOUSE COFFEE - BUSINESS REPORT"
ws_summary['A1'].font = Font(bold=True, size=18, color="4A3728")
ws_summary['A1'].alignment = Alignment(horizontal='center')

ws_summary.merge_cells('A2:F2')
ws_summary['A2'] = f"Report Period: October 18 - December 17, 2024 (60 Days)"
ws_summary['A2'].font = Font(size=12, italic=True)
ws_summary['A2'].alignment = Alignment(horizontal='center')

# Supplier Summary Section
row = 4
ws_summary.merge_cells(f'A{row}:C{row}')
ws_summary[f'A{row}'] = "SUPPLIER SPENDING"
apply_subheader_style(ws_summary[f'A{row}'])
apply_subheader_style(ws_summary[f'B{row}'])
apply_subheader_style(ws_summary[f'C{row}'])

row += 1
supplier_stats = [
    ("Total Purchase Orders", 8),
    ("Green Coffee Beans", 52101.70),
    ("Packaging Materials", 7660.00),
    ("Equipment/Parts", 1371.00),
    ("Shipping & Logistics", 11575.00),
    ("Import Duties", 2529.45),
    ("TOTAL SUPPLIER SPEND", 61563.95),
]
for label, value in supplier_stats:
    ws_summary[f'A{row}'] = label
    ws_summary[f'A{row}'].border = thin_border
    if isinstance(value, float):
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].number_format = currency_format
    else:
        ws_summary[f'B{row}'] = value
    ws_summary[f'B{row}'].border = thin_border
    if label == "TOTAL SUPPLIER SPEND":
        ws_summary[f'A{row}'].font = Font(bold=True)
        ws_summary[f'B{row}'].font = Font(bold=True)
    row += 1

# Customer Sales Section
row += 1
ws_summary.merge_cells(f'A{row}:C{row}')
ws_summary[f'A{row}'] = "CUSTOMER SALES"
apply_subheader_style(ws_summary[f'A{row}'])
apply_subheader_style(ws_summary[f'B{row}'])
apply_subheader_style(ws_summary[f'C{row}'])

row += 1
customer_stats = [
    ("Total Orders", 24),
    ("Total Revenue", 4108.62),
    ("Average Order Value", 171.19),
    ("Orders Delivered", "20 (83%)"),
    ("Orders In Transit", "2 (8%)"),
    ("Orders Processing/Pending", "2 (8%)"),
    ("Discounts Given", 363.69),
]
for label, value in customer_stats:
    ws_summary[f'A{row}'] = label
    ws_summary[f'A{row}'].border = thin_border
    if isinstance(value, float):
        ws_summary[f'B{row}'] = value
        ws_summary[f'B{row}'].number_format = currency_format
    else:
        ws_summary[f'B{row}'] = value
    ws_summary[f'B{row}'].border = thin_border
    row += 1

# Top Products Section
row += 1
ws_summary.merge_cells(f'D4:F4')
ws_summary['D4'] = "TOP SELLING PRODUCTS"
apply_subheader_style(ws_summary['D4'])
apply_subheader_style(ws_summary['E4'])
apply_subheader_style(ws_summary['F4'])

top_products = [
    ("Hair Bender", 8, "$407.60"),
    ("Holler Mountain", 5, "$259.95"),
    ("Ethiopia Mordecofe", 4, "$223.96"),
    ("Guatemala Injerto", 3, "$164.97"),
    ("Two Bag Subscription", 3, "$230.97"),
]
row_p = 5
for product, orders, revenue in top_products:
    ws_summary[f'D{row_p}'] = product
    ws_summary[f'D{row_p}'].border = thin_border
    ws_summary[f'E{row_p}'] = f"{orders} orders"
    ws_summary[f'E{row_p}'].border = thin_border
    ws_summary[f'F{row_p}'] = revenue
    ws_summary[f'F{row_p}'].border = thin_border
    row_p += 1

# Discount Code Usage
ws_summary.merge_cells(f'D11:F11')
ws_summary['D11'] = "DISCOUNT CODE USAGE"
apply_subheader_style(ws_summary['D11'])
apply_subheader_style(ws_summary['E11'])
apply_subheader_style(ws_summary['F11'])

discount_codes = [
    ("WELCOME10", 5, "$87.56"),
    ("HOLIDAY15", 6, "$156.21"),
    ("SAVE30", 4, "$89.98"),
]
row_d = 12
for code, times, savings in discount_codes:
    ws_summary[f'D{row_d}'] = code
    ws_summary[f'D{row_d}'].border = thin_border
    ws_summary[f'E{row_d}'] = f"{times} uses"
    ws_summary[f'E{row_d}'].border = thin_border
    ws_summary[f'F{row_d}'] = savings
    ws_summary[f'F{row_d}'].border = thin_border
    row_d += 1

auto_fit_columns(ws_summary)

# ============================================
# SHEET 2: Suppliers
# ============================================
ws_suppliers = wb.create_sheet("Suppliers")

headers = ["Supplier ID", "Company Name", "Contact", "Email", "Phone", "Location", "Specialty"]
for col, header in enumerate(headers, 1):
    cell = ws_suppliers.cell(row=1, column=col, value=header)
    apply_header_style(cell)

for row, supplier in enumerate(data['suppliers'], 2):
    ws_suppliers.cell(row=row, column=1, value=supplier['id']).border = thin_border
    ws_suppliers.cell(row=row, column=2, value=supplier['name']).border = thin_border
    ws_suppliers.cell(row=row, column=3, value=supplier['contact']).border = thin_border
    ws_suppliers.cell(row=row, column=4, value=supplier['email']).border = thin_border
    ws_suppliers.cell(row=row, column=5, value=supplier['phone']).border = thin_border
    ws_suppliers.cell(row=row, column=6, value=supplier['address']).border = thin_border
    ws_suppliers.cell(row=row, column=7, value=supplier['specialty']).border = thin_border

auto_fit_columns(ws_suppliers)

# ============================================
# SHEET 3: Supplier Purchase Orders
# ============================================
ws_po = wb.create_sheet("Purchase Orders")

headers = ["PO Number", "Supplier", "Order Date", "Expected Delivery", "Actual Delivery", 
           "Status", "Tracking Number", "Carrier", "Items", "Subtotal", "Shipping", 
           "Duties", "Total", "Payment Status"]

for col, header in enumerate(headers, 1):
    cell = ws_po.cell(row=1, column=col, value=header)
    apply_header_style(cell)

for row, po in enumerate(data['supplier_purchase_orders'], 2):
    items_str = "; ".join([f"{item['description']} ({item.get('quantity_kg', item.get('quantity'))})" for item in po['items']])
    
    ws_po.cell(row=row, column=1, value=po['po_number']).border = thin_border
    ws_po.cell(row=row, column=2, value=po['supplier_name']).border = thin_border
    ws_po.cell(row=row, column=3, value=po['order_date']).border = thin_border
    ws_po.cell(row=row, column=4, value=po['expected_delivery']).border = thin_border
    ws_po.cell(row=row, column=5, value=po['actual_delivery'] or "—").border = thin_border
    
    status_cell = ws_po.cell(row=row, column=6, value=po['status'])
    status_cell.border = thin_border
    if po['status'] == 'Delivered':
        status_cell.fill = delivered_fill
    elif po['status'] == 'In Transit':
        status_cell.fill = transit_fill
    
    ws_po.cell(row=row, column=7, value=po['tracking_number']).border = thin_border
    ws_po.cell(row=row, column=8, value=po['carrier']).border = thin_border
    ws_po.cell(row=row, column=9, value=items_str).border = thin_border
    
    subtotal_cell = ws_po.cell(row=row, column=10, value=po['subtotal'])
    subtotal_cell.number_format = currency_format
    subtotal_cell.border = thin_border
    
    shipping_cell = ws_po.cell(row=row, column=11, value=po['shipping'])
    shipping_cell.number_format = currency_format
    shipping_cell.border = thin_border
    
    duties_cell = ws_po.cell(row=row, column=12, value=po['import_duties'])
    duties_cell.number_format = currency_format
    duties_cell.border = thin_border
    
    total_cell = ws_po.cell(row=row, column=13, value=po['total'])
    total_cell.number_format = currency_format
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border
    
    ws_po.cell(row=row, column=14, value=po['payment_status']).border = thin_border

auto_fit_columns(ws_po)

# ============================================
# SHEET 4: Customer Orders
# ============================================
ws_orders = wb.create_sheet("Customer Orders")

headers = ["Order ID", "Order Date", "Ship Date", "Delivery Date", "Status", 
           "Tracking Number", "Customer Name", "Email", "City", "State", 
           "Items", "Subtotal", "Discount Code", "Discount", "Shipping", "Tax", "Total"]

for col, header in enumerate(headers, 1):
    cell = ws_orders.cell(row=1, column=col, value=header)
    apply_header_style(cell)

for row, order in enumerate(data['customer_orders'], 2):
    items_str = "; ".join([f"{item['product']} x{item['quantity']}" for item in order['items']])
    
    ws_orders.cell(row=row, column=1, value=order['order_id']).border = thin_border
    ws_orders.cell(row=row, column=2, value=order['order_date']).border = thin_border
    ws_orders.cell(row=row, column=3, value=order['ship_date'] or "—").border = thin_border
    ws_orders.cell(row=row, column=4, value=order['delivery_date'] or "—").border = thin_border
    
    status_cell = ws_orders.cell(row=row, column=5, value=order['status'])
    status_cell.border = thin_border
    if order['status'] == 'Delivered':
        status_cell.fill = delivered_fill
    elif order['status'] == 'In Transit':
        status_cell.fill = transit_fill
    elif order['status'] in ['Processing', 'Pending']:
        status_cell.fill = pending_fill
    
    ws_orders.cell(row=row, column=6, value=order['tracking_number'] or "—").border = thin_border
    ws_orders.cell(row=row, column=7, value=order['customer']['name']).border = thin_border
    ws_orders.cell(row=row, column=8, value=order['customer']['email']).border = thin_border
    ws_orders.cell(row=row, column=9, value=order['customer']['city']).border = thin_border
    ws_orders.cell(row=row, column=10, value=order['customer']['state']).border = thin_border
    ws_orders.cell(row=row, column=11, value=items_str).border = thin_border
    
    subtotal_cell = ws_orders.cell(row=row, column=12, value=order['subtotal'])
    subtotal_cell.number_format = currency_format
    subtotal_cell.border = thin_border
    
    ws_orders.cell(row=row, column=13, value=order['discount_code'] or "—").border = thin_border
    
    discount_cell = ws_orders.cell(row=row, column=14, value=order['discount_amount'])
    discount_cell.number_format = currency_format
    discount_cell.border = thin_border
    
    shipping_cell = ws_orders.cell(row=row, column=15, value=order['shipping'])
    shipping_cell.number_format = currency_format
    shipping_cell.border = thin_border
    
    tax_cell = ws_orders.cell(row=row, column=16, value=order['tax'])
    tax_cell.number_format = currency_format
    tax_cell.border = thin_border
    
    total_cell = ws_orders.cell(row=row, column=17, value=order['total'])
    total_cell.number_format = currency_format
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border

auto_fit_columns(ws_orders)

# ============================================
# SHEET 5: Inventory (Green Coffee)
# ============================================
ws_inventory = wb.create_sheet("Green Coffee Inventory")

headers = ["Origin", "Product", "Quantity (kg)", "Cost/kg", "Total Cost", "Supplier", "PO Number", "Received Date"]
for col, header in enumerate(headers, 1):
    cell = ws_inventory.cell(row=1, column=col, value=header)
    apply_header_style(cell)

inventory_data = [
    ("Colombia", "Colombia El Jordan", 500, 8.45, 4225.00, "Caravela Coffee Trading", "PO-2024-0847", "Nov 6, 2024"),
    ("Colombia", "Colombia Huila", 300, 7.90, 2370.00, "Caravela Coffee Trading", "PO-2024-0847", "Nov 6, 2024"),
    ("Ethiopia", "Ethiopia Mordecofe", 400, 11.20, 4480.00, "Ethiopian Yirgacheffe Exporters", "PO-2024-0863", "Nov 20, 2024"),
    ("Ethiopia", "Ethiopia Duromina", 350, 12.50, 4375.00, "Ethiopian Yirgacheffe Exporters", "PO-2024-0863", "Nov 20, 2024"),
    ("Ethiopia", "Ethiopia Suke Quto", 250, 14.80, 3700.00, "Ethiopian Yirgacheffe Exporters", "PO-2024-0863", "Nov 20, 2024"),
    ("Indonesia", "Indonesia Beis Penantan", 450, 9.75, 4387.50, "Sumatra Prima Cooperative", "PO-2024-0912", "Dec 3, 2024"),
    ("Indonesia", "Sumatra Mandheling", 300, 8.90, 2670.00, "Sumatra Prima Cooperative", "PO-2024-0912", "Dec 3, 2024"),
    ("Guatemala", "Guatemala Injerto", 400, 9.20, 3680.00, "Caravela Coffee Trading", "PO-2024-0934", "Dec 9, 2024"),
    ("Costa Rica", "Costa Rica Bella Vista", 350, 10.50, 3675.00, "Caravela Coffee Trading", "PO-2024-0934", "Dec 9, 2024"),
    ("Ethiopia", "Ethiopia Mordecofe (New Crop)", 600, 11.50, 6900.00, "Ethiopian Yirgacheffe Exporters", "PO-2024-0989", "In Transit"),
]

for row, inv in enumerate(inventory_data, 2):
    for col, value in enumerate(inv, 1):
        cell = ws_inventory.cell(row=row, column=col, value=value)
        cell.border = thin_border
        if col == 4:  # Cost/kg
            cell.number_format = currency_format
        if col == 5:  # Total cost
            cell.number_format = currency_format
            cell.font = Font(bold=True)

# Totals row
total_row = len(inventory_data) + 2
ws_inventory.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
ws_inventory.cell(row=total_row, column=3, value=3900).font = Font(bold=True)
ws_inventory.cell(row=total_row, column=3).border = thin_border
total_cost_cell = ws_inventory.cell(row=total_row, column=5, value=40462.50)
total_cost_cell.number_format = currency_format
total_cost_cell.font = Font(bold=True)
total_cost_cell.border = thin_border

auto_fit_columns(ws_inventory)

# ============================================
# Save the workbook
# ============================================
output_file = "Grainhouse_Coffee_Business_Report.xlsx"
wb.save(output_file)
print(f"\n[SUCCESS] Excel report generated: {output_file}")
print("\nSheets included:")
print("  1. Dashboard - Summary statistics")
print("  2. Suppliers - Vendor contact information")
print("  3. Purchase Orders - Supplier orders with tracking")
print("  4. Customer Orders - Sales with tracking numbers")
print("  5. Green Coffee Inventory - Coffee bean inventory")

